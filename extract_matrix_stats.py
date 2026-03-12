#!/usr/bin/env python3
"""
Extract summary statistics from deepTools computeMatrix output and create Excel file.
Combines sorted region BED files with signal statistics.
"""

import subprocess
import sys

# Check and install required packages
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    import openpyxl

import pandas as pd
import numpy as np
import gzip
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_deeptools_matrix(matrix_file):
    """Parse deepTools matrix file and extract signal data."""
    data = []
    with gzip.open(matrix_file, 'rt') as f:
        for line in f:
            if not line.startswith('@'):
                data.append(line.strip().split('\t'))
    
    return data

def extract_statistics(matrix_file, bed_file, output_excel):
    """Extract statistics and create Excel file."""
    
    # Parse matrix file - it already contains region info in first 6 columns
    matrix_data = parse_deeptools_matrix(matrix_file)
    
    print(f"Total regions in matrix: {len(matrix_data)}")
    
    # Extract signal values from matrix (columns 6 onwards contain signal)
    results = []
    for row in matrix_data:
        # First 6 columns are: chr, start, end, name, score, strand
        region_info = row[:6]
        signal_values = [float(x) if x != 'nan' else np.nan for x in row[6:]]
        
        # Calculate statistics
        signal_array = np.array(signal_values)
        signal_array = signal_array[~np.isnan(signal_array)]
        
        if len(signal_array) > 0:
            mean_signal = np.mean(signal_array)
            median_signal = np.median(signal_array)
            max_signal = np.max(signal_array)
            min_signal = np.min(signal_array)
            sum_signal = np.sum(signal_array)
        else:
            mean_signal = median_signal = max_signal = min_signal = sum_signal = 0
        
        results.append({
            'chr': region_info[0],
            'start': int(region_info[1]),
            'end': int(region_info[2]),
            'name': region_info[3],
            'score': region_info[4],
            'strand': region_info[5],
            'mean_signal': mean_signal,
            'median_signal': median_signal,
            'max_signal': max_signal,
            'min_signal': min_signal,
            'sum_signal': sum_signal
        })
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = 'Gene_Signal_Stats'
    
    # Headers
    headers = ['Chr', 'Start', 'End', 'Gene', 'Score', 'Strand', 
               'Mean Signal', 'Median Signal', 'Max Signal', 'Min Signal', 'Sum Signal']
    
    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, row_data in df.iterrows():
        ws.cell(row=row_idx+2, column=1, value=row_data['chr'])
        ws.cell(row=row_idx+2, column=2, value=row_data['start'])
        ws.cell(row=row_idx+2, column=3, value=row_data['end'])
        ws.cell(row=row_idx+2, column=4, value=row_data['name'])
        ws.cell(row=row_idx+2, column=5, value=row_data['score'])
        ws.cell(row=row_idx+2, column=6, value=row_data['strand'])
        ws.cell(row=row_idx+2, column=7, value=row_data['mean_signal'])
        ws.cell(row=row_idx+2, column=8, value=row_data['median_signal'])
        ws.cell(row=row_idx+2, column=9, value=row_data['max_signal'])
        ws.cell(row=row_idx+2, column=10, value=row_data['min_signal'])
        ws.cell(row=row_idx+2, column=11, value=row_data['sum_signal'])
    
    # Format signal columns with 2 decimal places
    for row in range(2, len(df) + 2):
        for col in range(7, 12):
            ws.cell(row=row, column=col).number_format = '0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    
    # Save
    wb.save(output_excel)
    print(f"Excel file created: {output_excel}")
    print(f"Total genes processed: {len(df)}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python extract_matrix_stats.py <matrix.gz> <output.xlsx>")
        print("\nExample:")
        print("  python extract_matrix_stats.py mute_p_gene_matrix.gz mute_p_signal_stats.xlsx")
        sys.exit(1)
    
    matrix_file = sys.argv[1]
    output_excel = sys.argv[2]
    
    extract_statistics(matrix_file, None, output_excel)